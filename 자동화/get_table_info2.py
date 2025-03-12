import openpyxl
import pandas 
from hdbcli import dbapi
import os
import re
import time


# 시작 시간 기록
start_time = time.perf_counter()

#def clean_illegal_chars(value):
#    if isinstance(value, str):
#        return re.sub(r'[\x0[1-0]', '', value)  # 문자열인 경우, 유효하지 않은 문자 제거
#    return value  # 문자열이 아닌 경우, 변환 없이 그대로 반환

# 현재 파일의 전체 경로
file_path = __file__

# 파일이 위치한 디렉토리 경로
directory_path = os.path.dirname(file_path)
directory_path = directory_path + '\\'

sample_file_name = 'table_profile.xlsx'
sample_file_name = directory_path + sample_file_name

sheet_name_01 = '01_CONNECTION_INFO'
sheet_name_02 = '02_DATA_TYPE_LIST'
sheet_name_03 = '03_SQL_LIST'
sheet_name_04 = '04_RESULT_TABLE'

wb=openpyxl.load_workbook(sample_file_name)

ws01 = wb[sheet_name_01]
ws02 = wb[sheet_name_02]
ws03 = wb[sheet_name_03]
ws04 = wb[sheet_name_04]

ws01_max_row = ws01.max_row
ws02_max_row = ws02.max_row
ws03_max_row = ws03.max_row
ws04_max_row = ws04.max_row


ws01_max_column = ws01.max_column
ws02_max_column = ws02.max_column
ws03_max_column = ws03.max_column
ws04_max_column = ws04.max_column


for x01 in range(1, ws01_max_row+1):
    if ws01.cell(row=x01, column=1).value == 1:
        if ws01.cell(row=x01, column=2).value == "address":
            hana_address=ws01.cell(row=x01, column=3).value
    elif ws01.cell(row=x01, column=1).value == 2:
        if ws01.cell(row=x01, column=2).value == "port":
            hana_port=ws01.cell(row=x01, column=3).value
    elif ws01.cell(row=x01, column=1).value == 3:
        if ws01.cell(row=x01, column=2).value == "user":
            hana_user=ws01.cell(row=x01, column=3).value
    elif ws01.cell(row=x01, column=1).value == 4:
        if ws01.cell(row=x01, column=2).value == "password":
            hana_password=ws01.cell(row=x01, column=3).value

# SAP HANA에 연결
conn01 = dbapi.connect(address=hana_address, port=hana_port, user=hana_user, password=hana_password)


# 실행할 SQL 쿼리
for x03 in range(1, ws03_max_row):
    if ws03.cell(row=x03, column=1).value == 9999:
        if ws03.cell(row=x03, column=2).value == "hana_meta":
            sql01=ws03.cell(row=x03, column=3).value
            
cur01 = conn01.cursor()

# 쿼리 실행
cur01.execute(sql01)
sql01=""
# 데이터 추가 시작 위치를 첫 번째 행으로 설정
start_row = 1

# 컬럼 헤더를 엑셀 첫 번째 행에 추가 (기존 데이터를 덮어쓰고 싶지 않다면 이 부분을 생략)
columns = [description[0] for description in cur01.description]
for col, header in enumerate(columns, start=1):
    ws04.cell(row=start_row, column=col, value=header)

# SQL 쿼리 결과를 엑셀 파일에 작성 (첫 번째 데이터 행을 다음 행부터 시작)
for row_idx, row in enumerate(cur01.fetchall(), start=start_row + 1):
    for col_idx, value in enumerate(row, start=1):
        ws04.cell(row=row_idx, column=col_idx, value=value)

# 컬럼 헤더
for i04 in range(4, ws02_max_column+1):
    ws04.cell(row=1, column=i04+8).value = ws02.cell(row=1, column=i04).value

for x04 in range(2, ws04_max_row):
    
    for x02 in range(2, ws02_max_row):
        if ws02.cell(row=x02, column=3).value == ws04.cell(row=x04, column=7).value:
            for y02 in range(4, ws02_max_column):
                #대용량테이블 부분 제외부분 추가
                if ws02.cell(row=x02, column=y02).value == "O":
                        if ws03.cell(row=y02-1, column=2).value=="mode_count" or \
                           ws03.cell(row=y02-1, column=2).value=="q1" or \
                           ws03.cell(row=y02-1, column=2).value=="q2" or \
                           ws03.cell(row=y02-1, column=2).value=="q3" or \
                           ws03.cell(row=y02-1, column=2).value=="q4" or \
                           ws03.cell(row=y02-1, column=2).value=="IQR" or \
                           ws03.cell(row=y02-1, column=2).value=="q1-(1.5*IQR)" or \
                           ws03.cell(row=y02-1, column=2).value=="q3+(1.5*IQR)" or \
                           sql01=""
                            
                        else:
                            if ws02.cell(row=x02, column=2).value == "INT" and ws03.cell(row=y02-1, column=2).value=="q0":
                                sql01=ws03.cell(row=y02-1, column=4).value.replace("Table_Name", ws04.cell(row=x04, column=3).value + '.' + ws04.cell(row=x04, column=4).value)
                                sql01=sql01.replace("Column_Name", ws04.cell(row=x04, column=6).value)
                            else:  
                                sql01=ws03.cell(row=y02-1, column=3).value.replace("Table_Name", ws04.cell(row=x04, column=3).value + '.' + ws04.cell(row=x04, column=4).value)
                                sql01=sql01.replace("Column_Name", ws04.cell(row=x04, column=6).value)

                            #예약어 ORDER에 대해서 "ORDER"로 변경
                            if ws04.cell(row=x04, column=6).value=="ORDER":
                                sql01=sql01.replace("ORDER", "\"ORDER\"")
                                sql01=sql01.replace("\"ORDER\" BY", "ORDER BY")
                            
                            #예약어 START에 대해서 "START"로 변경    
                            if ws04.cell(row=x04, column=6).value=="START":
                                sql01=sql01.replace("START", "\"START\"")     
                            
                            #예약어 END에 대해서 "END"로 변경    
                            if ws04.cell(row=x04, column=6).value=="END":
                                sql01=sql01.replace("END", "\"END\"")
                                sql01=sql01.replace("NULL \"END\"", "NULL END")

                            #예약어 UNION에 대해서 "UNION"로 변경    
                            if ws04.cell(row=x04, column=6).value=="UNION":
                                sql01=sql01.replace("UNION", "\"UNION\"")
 
                            # 쿼리 실행
                            
                            try:

                                print(sql01)

                                cur01.execute(sql01)
                                rows01 = cur01.fetchall()
                                             
                           
                            #else:
                                # 쿼리 결과를 엑셀 파일에 쓰기
                                start_row01 = 1
                                start_column01 = 1
                                for row_idx01, row01 in enumerate(rows01, start=start_row01):
                                    for col_idx01, value01 in enumerate(row01, start=start_column01):
                                        #ws.cell(row=row_idx, column=col_idx, value=value)
                                        ws04.cell(row=x04, column=y02+8+(col_idx01-1), value=value01)

                            except Exception as e:
                                ERR_MSG="Error: " + str(e)
                                ws04.cell(row=x04, column=y02+8, value=ERR_MSG)


# 시트 저장
wb.save(sample_file_name)

cur01.close()
conn01.close()

# 종료 시간 기록
end_time = time.perf_counter()

# 수행 시간 계산 및 출력
print(f"Time taken: {end_time - start_time} seconds")